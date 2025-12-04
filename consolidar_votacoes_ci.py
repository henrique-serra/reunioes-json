#!/usr/bin/env python3
"""
Script para consolidar informações de votações da Comissão de Infraestrutura (CI)
a partir de arquivos JSON de reuniões do Senado Federal.
"""

import json
import glob
import os
from pathlib import Path
import pandas as pd
from typing import List, Dict, Any
from datetime import datetime
import requests
import sys


def solicitar_ano() -> str:
    """
    Solicita o ano ao usuário via input.

    Returns:
        Ano como string
    """
    while True:
        ano = input("Digite o ano que deseja analisar (ex: 2025): ").strip()

        if not ano.isdigit():
            print("Erro: Digite apenas números.")
            continue

        if len(ano) != 4:
            print("Erro: O ano deve ter 4 dígitos.")
            continue

        ano_int = int(ano)
        if ano_int < 2000 or ano_int > 2100:
            print("Erro: Digite um ano válido entre 2000 e 2100.")
            continue

        return ano


def verificar_arquivos_existentes(pasta: str, ano: str) -> tuple[bool, bool]:
    """
    Verifica se os arquivos JSON do ano especificado existem na pasta.

    Args:
        pasta: Caminho da pasta a verificar
        ano: Ano a verificar

    Returns:
        Tupla (tem_reunioes, tem_colegiados) indicando quais arquivos existem
    """
    # Verificar arquivos de reuniões (12 meses)
    arquivos_reunioes = []
    for mes in range(1, 13):
        mes_str = f"{mes:02d}"
        arquivo = os.path.join(pasta, f"{ano}{mes_str}-reunioes.json")
        if os.path.exists(arquivo):
            arquivos_reunioes.append(arquivo)

    tem_reunioes = len(arquivos_reunioes) == 12

    # Verificar lista de colegiados
    arquivo_colegiados = os.path.join(pasta, "lista_colegiados.json")
    tem_colegiados = os.path.exists(arquivo_colegiados)

    return tem_reunioes, tem_colegiados


def baixar_arquivo_reunioes(ano: str, mes: str, pasta: str) -> bool:
    """
    Baixa um arquivo de reuniões de um mês específico.

    Args:
        ano: Ano a baixar
        mes: Mês a baixar (formato: "01", "02", etc)
        pasta: Pasta onde salvar o arquivo

    Returns:
        True se sucesso, False se falhou
    """
    url = f"https://legis.senado.leg.br/dadosabertos/comissao/agenda/mes/{ano}{mes}.json"
    filename = os.path.join(pasta, f"{ano}{mes}-reunioes.json")

    try:
        print(f"  Baixando {ano}{mes}-reunioes.json...", end=" ")
        response = requests.get(url, timeout=30)
        response.raise_for_status()

        data = response.json()

        with open(filename, "w", encoding="utf-8") as file:
            json.dump(data, file, indent=2, ensure_ascii=False)

        print("OK")
        return True

    except requests.exceptions.RequestException as e:
        print(f"ERRO: {e}")
        return False
    except json.JSONDecodeError as e:
        print(f"ERRO ao decodificar JSON: {e}")
        return False
    except Exception as e:
        print(f"ERRO inesperado: {e}")
        return False


def baixar_lista_colegiados(pasta: str) -> bool:
    """
    Baixa o arquivo de lista de colegiados.

    Args:
        pasta: Pasta onde salvar o arquivo

    Returns:
        True se sucesso, False se falhou
    """
    url = "https://legis.senado.leg.br/dadosabertos/comissao/lista/colegiados.json"
    filename = os.path.join(pasta, "lista_colegiados.json")

    try:
        print("  Baixando lista_colegiados.json...", end=" ")
        response = requests.get(url, timeout=30)
        response.raise_for_status()

        data = response.json()

        with open(filename, "w", encoding="utf-8") as file:
            json.dump(data, file, indent=2, ensure_ascii=False)

        print("OK")
        return True

    except requests.exceptions.RequestException as e:
        print(f"ERRO: {e}")
        return False
    except json.JSONDecodeError as e:
        print(f"ERRO ao decodificar JSON: {e}")
        return False
    except Exception as e:
        print(f"ERRO inesperado: {e}")
        return False


def baixar_arquivos_ano(ano: str, pasta: str) -> bool:
    """
    Baixa todos os arquivos necessários para um ano específico.

    Args:
        ano: Ano a baixar
        pasta: Pasta onde salvar os arquivos

    Returns:
        True se todos os downloads tiveram sucesso, False se algum falhou
    """
    print(f"\n{'='*70}")
    print(f"BAIXANDO ARQUIVOS DO ANO {ano}")
    print(f"{'='*70}\n")

    sucesso_total = True

    # Baixar lista de colegiados
    if not baixar_lista_colegiados(pasta):
        sucesso_total = False

    # Baixar arquivos de reuniões para cada mês
    meses = [f"{m:02d}" for m in range(1, 13)]
    for mes in meses:
        if not baixar_arquivo_reunioes(ano, mes, pasta):
            sucesso_total = False

    print(f"\n{'='*70}")
    if sucesso_total:
        print("DOWNLOAD CONCLUÍDO COM SUCESSO")
    else:
        print("DOWNLOAD CONCLUÍDO COM ALGUNS ERROS")
    print(f"{'='*70}\n")

    return sucesso_total


def classificar_tipo_requerimento(ementa: str) -> str:
    """
    Classifica o tipo de requerimento baseado na ementa.

    Args:
        ementa: Texto da ementa do requerimento

    Returns:
        Tipo do requerimento classificado
    """
    if not ementa:
        return ''

    ementa_lower = ementa.lower()

    # Verificar aditamento primeiro (pode ser aditamento de audiência)
    if 'aditamento' in ementa_lower:
        if 'audiência pública' in ementa_lower or 'audiencia publica' in ementa_lower:
            return 'Aditamento de Audiência Pública'
        return 'Aditamento'

    # Verificar audiência pública
    if 'audiência pública' in ementa_lower or 'audiencia publica' in ementa_lower:
        return 'Audiência Pública'

    # Verificar outros tipos comuns
    if 'retirada' in ementa_lower and 'emenda' in ementa_lower:
        return 'Retirada de Emenda'

    if 'informação' in ementa_lower or 'informações' in ementa_lower:
        return 'Pedido de Informação'
    
    if 'diligência' in ementa_lower or 'diligências' in ementa_lower:
        return 'Diligência'

    if 'convocação' in ementa_lower or 'convocar' in ementa_lower:
        return 'Convocação'

    if 'urgência' in ementa_lower or 'urgencia' in ementa_lower:
        return 'Urgência'

    if 'voto de' in ementa_lower:
        return 'Voto de Aplauso/Pesar/Congratulações'

    # Tipo genérico
    return 'Requerimento (Outros)'


def extrair_votacoes_ci(arquivo_json: str) -> List[Dict[str, Any]]:
    """
    Extrai votações da Comissão de Infraestrutura de um arquivo JSON.

    Args:
        arquivo_json: Caminho para o arquivo JSON

    Returns:
        Lista de dicionários com informações das votações
    """
    votacoes = []

    try:
        with open(arquivo_json, 'r', encoding='utf-8') as f:
            dados = json.load(f)

        # Navegar até as reuniões
        reunioes = dados.get('AgendaReuniao', {}).get('reunioes', {})

        # Se reunioes for None, pular o arquivo
        if not reunioes:
            return votacoes

        reuniao_list = reunioes.get('reuniao', [])

        # Garantir que reuniao_list seja uma lista
        if isinstance(reuniao_list, dict):
            reuniao_list = [reuniao_list]

        for reuniao in reuniao_list:
            # Verificar se reunião é válida
            if not reuniao or not isinstance(reuniao, dict):
                continue

            # Filtrar apenas reuniões da Comissão de Infraestrutura (sigla == 'CI')
            colegiados = reuniao.get('colegiados')
            if not colegiados:
                continue

            # Tratar caso colegiados seja lista (múltiplos colegiados)
            if isinstance(colegiados, list):
                siglas = [c.get('sigla') for c in colegiados if isinstance(c, dict)]
                if 'CI' not in siglas:
                    continue
            elif isinstance(colegiados, dict):
                if colegiados.get('sigla') != 'CI':
                    continue
            else:
                continue

            # Extrair informações da reunião
            titulo_reuniao = reuniao.get('titulo', '')
            realizada = reuniao.get('realizada', '')
            data_inicio_raw = reuniao.get('dataInicio', '')

            # Converter data_inicio para objeto datetime (não string)
            data_inicio = None
            if data_inicio_raw:
                try:
                    # A data vem no formato ISO: 2025-09-24T09:00:00.000
                    data_inicio = datetime.fromisoformat(data_inicio_raw.replace('Z', '+00:00'))
                except (ValueError, AttributeError):
                    data_inicio = None

            # Processar as partes da reunião
            partes = reuniao.get('partes')

            # Se partes for None ou vazio, pular
            if not partes:
                continue

            # Verificar se partes é uma lista ou um objeto único
            if isinstance(partes, dict):
                partes = [partes]
            elif not isinstance(partes, list):
                continue

            # Filtrar partes deliberativas
            for parte in partes:
                if not isinstance(parte, dict):
                    continue

                descricao_tipo = parte.get('descricaoTipo', '')
                if descricao_tipo != 'Deliberativa':
                    continue

                # Processar os itens da parte deliberativa
                itens = parte.get('itens')

                # Se itens for None ou vazio, pular
                if not itens:
                    continue

                # Garantir que itens seja uma lista
                if isinstance(itens, dict):
                    itens = [itens]
                elif not isinstance(itens, list):
                    continue

                for item in itens:
                    if not isinstance(item, dict):
                        continue

                    # Extrair informações do item
                    descricao_item = item.get('descricao', '')
                    situacao = item.get('situacao', {})
                    descricao_tipo_resultado = situacao.get('descricaoTipoResultado', '') if isinstance(situacao, dict) else ''

                    # Extrair informações de requerimento (se existir)
                    # Pode estar em 'requerimento', 'materiaNaoAutonoma' ou 'doma'
                    requerimento = item.get('requerimento') or item.get('materiaNaoAutonoma') or item.get('doma')
                    ementa = ''
                    tipo_requerimento = ''
                    sigla_por_extenso = ''

                    # Verificar se é realmente um requerimento (tem "REQ" na descrição)
                    is_requerimento = 'REQ' in descricao_item.upper()

                    if requerimento and isinstance(requerimento, dict):
                        ementa = requerimento.get('ementa', '')
                        sigla_por_extenso = requerimento.get('siglaPorExtenso', '')

                        # Classificar tipo apenas se for requerimento
                        if is_requerimento and ementa:
                            tipo_requerimento = classificar_tipo_requerimento(ementa)
                        elif not is_requerimento:
                            # Se não é requerimento, deixar em branco
                            tipo_requerimento = '-'

                    # Verificar se a matéria foi apreciada
                    # A matéria foi apreciada se o resultado contiver os termos:
                    # aprovado, aprovada, rejeitado ou rejeitada (case-insensitive)
                    resultado_lower = descricao_tipo_resultado.lower()
                    foi_apreciada = any(termo in resultado_lower for termo in ['aprovado', 'aprovada', 'rejeitado', 'rejeitada'])
                    apreciado = 'Sim' if foi_apreciada else 'Não'

                    # Verificar se foi sabatina realizada
                    sabatina_realizada = 'Sim' if 'sabatina realizada' in resultado_lower else '-'

                    # Verificar se sabatina foi apreciada
                    if 'sabatina realizada com indicação aprovada' in resultado_lower or 'sabatina realizada com indicação rejeitada' in resultado_lower:
                        sabatina_apreciada = 'Sim'
                    elif 'sabatina realizada sem apreciação' in resultado_lower:
                        sabatina_apreciada = 'Não'
                    else:
                        sabatina_apreciada = '-'

                    # Verificar se sigla_por_extenso contém o termo "projeto" (case insensitive)
                    projeto = 'Sim' if sigla_por_extenso and 'projeto' in sigla_por_extenso.lower() else 'Não'

                    # Adicionar à lista de votações
                    votacao = {
                        'titulo_reuniao': titulo_reuniao,
                        'realizada': realizada,
                        'data_inicio': data_inicio,
                        'descricao_item': descricao_item,
                        'sigla_por_extenso': sigla_por_extenso,
                        'projeto': projeto,
                        'ementa': ementa,
                        'tipo_requerimento': tipo_requerimento,
                        'resultado': descricao_tipo_resultado,
                        'apreciado': apreciado,
                        'sabatina_realizada': sabatina_realizada,
                        'sabatina_apreciada': sabatina_apreciada,
                        'arquivo_origem': os.path.basename(arquivo_json)
                    }
                    votacoes.append(votacao)

    except Exception as e:
        print(f"Erro ao processar {arquivo_json}: {str(e)}")

    return votacoes


def extrair_audiencias_publicas_ci(arquivo_json: str) -> List[Dict[str, Any]]:
    """
    Extrai audiências públicas da Comissão de Infraestrutura de um arquivo JSON.

    Args:
        arquivo_json: Caminho para o arquivo JSON

    Returns:
        Lista de dicionários com informações das audiências públicas
    """
    audiencias = []

    try:
        with open(arquivo_json, 'r', encoding='utf-8') as f:
            dados = json.load(f)

        # Navegar até as reuniões
        reunioes = dados.get('AgendaReuniao', {}).get('reunioes', {})

        # Se reunioes for None, pular o arquivo
        if not reunioes:
            return audiencias

        reuniao_list = reunioes.get('reuniao', [])

        # Garantir que reuniao_list seja uma lista
        if isinstance(reuniao_list, dict):
            reuniao_list = [reuniao_list]

        for reuniao in reuniao_list:
            # Verificar se reunião é válida
            if not reuniao or not isinstance(reuniao, dict):
                continue

            # Filtrar apenas reuniões da Comissão de Infraestrutura (sigla == 'CI')
            colegiados = reuniao.get('colegiados')
            if not colegiados:
                continue

            # Tratar caso colegiados seja lista (múltiplos colegiados)
            if isinstance(colegiados, list):
                siglas = [c.get('sigla') for c in colegiados if isinstance(c, dict)]
                if 'CI' not in siglas:
                    continue
            elif isinstance(colegiados, dict):
                if colegiados.get('sigla') != 'CI':
                    continue
            else:
                continue

            # Extrair informações da reunião
            titulo_reuniao = reuniao.get('titulo', '')
            realizada = reuniao.get('realizada', '')
            data_inicio_raw = reuniao.get('dataInicio', '')

            # Converter data_inicio para objeto datetime
            data_inicio = None
            if data_inicio_raw:
                try:
                    data_inicio = datetime.fromisoformat(data_inicio_raw.replace('Z', '+00:00'))
                except (ValueError, AttributeError):
                    data_inicio = None

            # Processar as partes da reunião
            partes = reuniao.get('partes')

            # Se partes for None ou vazio, pular
            if not partes:
                continue

            # Verificar se partes é uma lista ou um objeto único
            if isinstance(partes, dict):
                partes = [partes]
            elif not isinstance(partes, list):
                continue

            # Procurar por partes de Audiência Pública
            for parte in partes:
                if not isinstance(parte, dict):
                    continue

                descricao_tipo = parte.get('descricaoTipo', '')

                # Verificar se é uma audiência pública
                if 'audiência' not in descricao_tipo.lower() and 'audiencia' not in descricao_tipo.lower():
                    continue

                # Extrair informações do evento
                evento = parte.get('evento')

                if not evento or not isinstance(evento, dict):
                    continue

                finalidade = evento.get('finalidade', '')
                convidados = evento.get('convidados', [])
                participantes = evento.get('participantes', [])

                # Garantir que sejam listas
                if isinstance(convidados, dict):
                    convidados = [convidados]
                if isinstance(participantes, dict):
                    participantes = [participantes]

                # Contar quantidade de convidados/participantes
                qtd_convidados = len(convidados) if convidados else 0
                qtd_participantes = len(participantes) if participantes else 0

                # Adicionar à lista de audiências
                audiencia = {
                    'titulo_reuniao': titulo_reuniao,
                    'data_inicio': data_inicio,
                    'realizada': realizada,
                    'finalidade': finalidade,
                    'qtd_convidados': qtd_convidados,
                    'qtd_participantes': qtd_participantes,
                    'arquivo_origem': os.path.basename(arquivo_json)
                }
                audiencias.append(audiencia)

    except Exception as e:
        print(f"Erro ao processar {arquivo_json}: {str(e)}")

    return audiencias


def extrair_partes_reuniao_ci(arquivo_json: str) -> List[Dict[str, Any]]:
    """
    Extrai informações sobre as partes das reuniões da Comissão de Infraestrutura.

    Args:
        arquivo_json: Caminho para o arquivo JSON

    Returns:
        Lista de dicionários com informações das partes
    """
    partes_info = []

    try:
        with open(arquivo_json, 'r', encoding='utf-8') as f:
            dados = json.load(f)

        # Navegar até as reuniões
        reunioes = dados.get('AgendaReuniao', {}).get('reunioes', {})

        # Se reunioes for None, pular o arquivo
        if not reunioes:
            return partes_info

        reuniao_list = reunioes.get('reuniao', [])

        # Garantir que reuniao_list seja uma lista
        if isinstance(reuniao_list, dict):
            reuniao_list = [reuniao_list]

        for reuniao in reuniao_list:
            # Verificar se reunião é válida
            if not reuniao or not isinstance(reuniao, dict):
                continue

            # Filtrar apenas reuniões da Comissão de Infraestrutura (sigla == 'CI')
            colegiados = reuniao.get('colegiados')
            if not colegiados:
                continue

            # Verificar se é reunião conjunta e extrair comissões participantes
            eh_reuniao_conjunta = False
            comissoes_participantes = []

            # Tratar caso colegiados seja lista (múltiplos colegiados = reunião conjunta)
            if isinstance(colegiados, list):
                siglas = [c.get('sigla') for c in colegiados if isinstance(c, dict)]
                if 'CI' not in siglas:
                    continue
                eh_reuniao_conjunta = len(colegiados) > 1
                # Pegar todas as siglas exceto CI
                comissoes_participantes = [c.get('sigla') for c in colegiados if isinstance(c, dict) and c.get('sigla') != 'CI']
            elif isinstance(colegiados, dict):
                if colegiados.get('sigla') != 'CI':
                    continue
                eh_reuniao_conjunta = False
            else:
                continue

            # Extrair informações da reunião
            titulo_reuniao = reuniao.get('titulo', '')
            realizada = reuniao.get('realizada', '')
            data_inicio_raw = reuniao.get('dataInicio', '')

            # Converter data_inicio para objeto datetime
            data_inicio = None
            if data_inicio_raw:
                try:
                    data_inicio = datetime.fromisoformat(data_inicio_raw.replace('Z', '+00:00'))
                except (ValueError, AttributeError):
                    data_inicio = None

            # Processar as partes da reunião
            partes = reuniao.get('partes')

            # Se partes for None ou vazio, pular
            if not partes:
                continue

            # Verificar se partes é uma lista ou um objeto único
            if isinstance(partes, dict):
                partes = [partes]
            elif not isinstance(partes, list):
                continue

            # Extrair informações de cada parte
            for parte in partes:
                if not isinstance(parte, dict):
                    continue

                descricao_tipo = parte.get('descricaoTipo', '')

                # Formatar comissões participantes
                if eh_reuniao_conjunta and comissoes_participantes:
                    comissoes_str = ', '.join(comissoes_participantes)
                else:
                    comissoes_str = '-'

                # Adicionar à lista
                parte_info = {
                    'titulo_reuniao': titulo_reuniao,
                    'realizada': realizada,
                    'data_inicio': data_inicio,
                    'tipo_parte': descricao_tipo,
                    'reuniao_conjunta': 'Sim' if eh_reuniao_conjunta else 'Não',
                    'comissoes_participantes': comissoes_str,
                    'arquivo_origem': os.path.basename(arquivo_json)
                }
                partes_info.append(parte_info)

    except Exception as e:
        print(f"Erro ao processar {arquivo_json}: {str(e)}")

    return partes_info


def gerar_resumo_anual(df_votacoes: pd.DataFrame, df_audiencias: pd.DataFrame, df_partes: pd.DataFrame) -> pd.DataFrame:
    """
    Gera um DataFrame com o resumo anual contendo todas as tabelas solicitadas.

    Args:
        df_votacoes: DataFrame com votações
        df_audiencias: DataFrame com audiências públicas
        df_partes: DataFrame com partes das reuniões

    Returns:
        DataFrame com o resumo formatado
    """
    resumo_linhas = []

    # ========== TABELA 1: Reuniões realizadas ==========
    resumo_linhas.append(['REUNIÕES REALIZADAS'])
    resumo_linhas.append(['Tipo de reunião', 'Quantidade'])

    if not df_partes.empty:
        # Filtrar apenas reuniões realizadas
        # Aceitar tanto 'true' (string do JSON) quanto True (booleano do Excel)
        df_partes_realizadas = df_partes[
            (df_partes['realizada'] == True) | (df_partes['realizada'] == 'true')
        ]

        # Contar por tipo de parte e ordenar por quantidade decrescente
        contagem_partes = df_partes_realizadas['tipo_parte'].value_counts().sort_values(ascending=False)

        for tipo, qtd in contagem_partes.items():
            resumo_linhas.append([tipo, qtd])

        resumo_linhas.append(['Total', contagem_partes.sum()])
    else:
        resumo_linhas.append(['Total', 0])

    # Linha em branco
    resumo_linhas.append([''])

    # ========== TABELA 2: Audiências públicas ==========
    resumo_linhas.append(['AUDIÊNCIAS PÚBLICAS'])
    resumo_linhas.append(['Quantidade de convidados', 'Quantidade de participantes'])

    if not df_audiencias.empty:
        # Filtrar apenas audiências realizadas
        # Aceitar tanto 'true' (string do JSON) quanto True (booleano do Excel)
        df_audiencias_realizadas = df_audiencias[
            (df_audiencias['realizada'] == True) | (df_audiencias['realizada'] == 'true')
        ]

        total_convidados = df_audiencias_realizadas['qtd_convidados'].sum()
        total_participantes = df_audiencias_realizadas['qtd_participantes'].sum()

        resumo_linhas.append([total_convidados, total_participantes])
    else:
        resumo_linhas.append([0, 0])

    # Linha em branco
    resumo_linhas.append([''])

    # ========== TABELA 3: Sabatinas ==========
    resumo_linhas.append(['SABATINAS'])
    resumo_linhas.append(['Sabatinas realizadas', 'Indicações apreciadas'])

    if not df_votacoes.empty:
        # Sabatinas realizadas (onde sabatina_realizada == 'Sim')
        sabatinas_realizadas = df_votacoes[df_votacoes['sabatina_realizada'] == 'Sim'].shape[0]

        # Indicações apreciadas (onde sabatina_apreciada == 'Sim')
        indicacoes_apreciadas = df_votacoes[df_votacoes['sabatina_apreciada'] == 'Sim'].shape[0]

        resumo_linhas.append([sabatinas_realizadas, indicacoes_apreciadas])
    else:
        resumo_linhas.append([0, 0])

    # Linha em branco
    resumo_linhas.append([''])

    # ========== TABELA 4: Projetos apreciados ==========
    resumo_linhas.append(['PROJETOS APRECIADOS'])
    resumo_linhas.append(['Tipo de projeto', 'Quantidade'])

    if not df_votacoes.empty:
        # Filtrar: apreciado == 'Sim' E projeto == 'Sim'
        df_projetos_apreciados = df_votacoes[
            (df_votacoes['apreciado'] == 'Sim') &
            (df_votacoes['projeto'] == 'Sim')
        ]

        if not df_projetos_apreciados.empty:
            # Contar por sigla_por_extenso e ordenar por quantidade decrescente
            contagem_projetos = df_projetos_apreciados['sigla_por_extenso'].value_counts().sort_values(ascending=False)

            for tipo, qtd in contagem_projetos.items():
                resumo_linhas.append([tipo, qtd])

            resumo_linhas.append(['Total', contagem_projetos.sum()])
        else:
            resumo_linhas.append(['Total', 0])
    else:
        resumo_linhas.append(['Total', 0])

    # Linha em branco
    resumo_linhas.append([''])

    # ========== TABELA 5: Requerimentos apreciados ==========
    resumo_linhas.append(['REQUERIMENTOS APRECIADOS'])
    resumo_linhas.append(['Tipo de requerimento', 'Quantidade'])

    if not df_votacoes.empty:
        # Filtrar: apreciado == 'Sim' E sigla_por_extenso contém 'Requerimento'
        df_req_apreciados = df_votacoes[
            (df_votacoes['apreciado'] == 'Sim') &
            (df_votacoes['sigla_por_extenso'].str.contains('Requerimento', case=False, na=False))
        ]

        if not df_req_apreciados.empty:
            # Contar por tipo_requerimento e ordenar por quantidade decrescente
            contagem_req = df_req_apreciados['tipo_requerimento'].value_counts().sort_values(ascending=False)

            for tipo, qtd in contagem_req.items():
                if tipo and tipo != '-':  # Ignorar vazios e '-'
                    resumo_linhas.append([tipo, qtd])

            resumo_linhas.append(['Total', contagem_req.sum()])
        else:
            resumo_linhas.append(['Total', 0])
    else:
        resumo_linhas.append(['Total', 0])

    # Linha em branco
    resumo_linhas.append([''])

    # ========== TABELA 6: Outras matérias apreciadas ==========
    resumo_linhas.append(['OUTRAS MATÉRIAS APRECIADAS'])
    resumo_linhas.append(['Tipo de matéria', 'Quantidade'])

    if not df_votacoes.empty:
        # Filtrar: apreciado == 'Sim' E projeto == 'Não' E sigla_por_extenso NÃO contém 'Requerimento'
        df_outras = df_votacoes[
            (df_votacoes['apreciado'] == 'Sim') &
            (df_votacoes['projeto'] == 'Não') &
            (~df_votacoes['sigla_por_extenso'].str.contains('Requerimento', case=False, na=False))
        ]

        if not df_outras.empty:
            # Contar por sigla_por_extenso e ordenar por quantidade decrescente
            contagem_outras = df_outras['sigla_por_extenso'].value_counts().sort_values(ascending=False)

            for tipo, qtd in contagem_outras.items():
                resumo_linhas.append([tipo, qtd])

            resumo_linhas.append(['Total', contagem_outras.sum()])
        else:
            resumo_linhas.append(['Total', 0])
    else:
        resumo_linhas.append(['Total', 0])

    # Linha em branco
    resumo_linhas.append([''])

    # ========== TABELA 7: Emendas ao orçamento (vazia) ==========
    resumo_linhas.append(['EMENDAS AO ORÇAMENTO'])
    resumo_linhas.append(['Tipo de matéria', 'Quantidade'])
    resumo_linhas.append(['PLOA', ''])
    resumo_linhas.append(['PLDO', ''])

    # Criar DataFrame com as linhas do resumo
    df_resumo = pd.DataFrame(resumo_linhas)

    return df_resumo


def gerar_planilha_consolidada(pasta_json: str, ano: str, arquivo_saida: str = None):
    """
    Processa todos os arquivos JSON de um ano específico e gera planilha consolidada.

    Args:
        pasta_json: Pasta contendo os arquivos JSON
        ano: Ano a processar
        arquivo_saida: Nome do arquivo Excel de saída (opcional)
    """
    # Definir nome do arquivo de saída se não fornecido
    if arquivo_saida is None:
        arquivo_saida = f'votacoes_ci_consolidadas_{ano}.xlsx'

    # Encontrar todos os arquivos JSON que começam com o ano especificado
    padrao_busca = os.path.join(pasta_json, f'{ano}*.json')
    arquivos_json = sorted(glob.glob(padrao_busca))

    if not arquivos_json:
        print(f"Nenhum arquivo JSON encontrado com o padrão '{ano}*.json' em {pasta_json}")
        return

    print(f"Encontrados {len(arquivos_json)} arquivos JSON:")
    for arquivo in arquivos_json:
        print(f"  - {os.path.basename(arquivo)}")

    # Processar todos os arquivos para votações
    todas_votacoes = []
    for arquivo in arquivos_json:
        print(f"\nProcessando votações de {os.path.basename(arquivo)}...")
        votacoes = extrair_votacoes_ci(arquivo)
        todas_votacoes.extend(votacoes)
        print(f"  Encontradas {len(votacoes)} votações")

    # Processar todos os arquivos para audiências públicas
    print("\n" + "="*70)
    print("PROCESSANDO AUDIÊNCIAS PÚBLICAS")
    print("="*70)
    todas_audiencias = []
    for arquivo in arquivos_json:
        print(f"\nProcessando audiências de {os.path.basename(arquivo)}...")
        audiencias = extrair_audiencias_publicas_ci(arquivo)
        todas_audiencias.extend(audiencias)
        print(f"  Encontradas {len(audiencias)} audiências públicas")

    # Processar todos os arquivos para partes das reuniões
    print("\n" + "="*70)
    print("PROCESSANDO PARTES DAS REUNIÕES")
    print("="*70)
    todas_partes = []
    for arquivo in arquivos_json:
        print(f"\nProcessando partes de {os.path.basename(arquivo)}...")
        partes = extrair_partes_reuniao_ci(arquivo)
        todas_partes.extend(partes)
        print(f"  Encontradas {len(partes)} partes")

    # Criar DataFrame de votações
    if todas_votacoes:
        df_votacoes = pd.DataFrame(todas_votacoes)
        df_votacoes = df_votacoes.sort_values('data_inicio', ascending=True)
    else:
        df_votacoes = pd.DataFrame()

    # Criar DataFrame de audiências públicas
    if todas_audiencias:
        df_audiencias = pd.DataFrame(todas_audiencias)
        df_audiencias = df_audiencias.sort_values('data_inicio', ascending=True)
    else:
        df_audiencias = pd.DataFrame()

    # Criar DataFrame de partes das reuniões
    if todas_partes:
        df_partes = pd.DataFrame(todas_partes)
        df_partes = df_partes.sort_values('data_inicio', ascending=True)
    else:
        df_partes = pd.DataFrame()

    # Gerar resumo anual
    print(f"\n{'='*70}")
    print("GERANDO RESUMO ANUAL")
    print(f"{'='*70}\n")
    df_resumo = gerar_resumo_anual(df_votacoes, df_audiencias, df_partes)

    # Salvar em Excel com múltiplas abas
    caminho_saida = os.path.join(pasta_json, arquivo_saida)

    with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
        # Aba de resumo anual (primeira aba)
        if not df_resumo.empty:
            df_resumo.to_excel(writer, index=False, sheet_name='Resumo Anual', header=False)

            # Aplicar formatação na aba Resumo Anual
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
            worksheet = writer.sheets['Resumo Anual']

            # Lista de títulos conhecidos (todas em maiúsculas)
            titulos = [
                'REUNIÕES REALIZADAS',
                'AUDIÊNCIAS PÚBLICAS',
                'SABATINAS',
                'PROJETOS APRECIADOS',
                'REQUERIMENTOS APRECIADOS',
                'OUTRAS MATÉRIAS APRECIADAS',
                'EMENDAS AO ORÇAMENTO'
            ]

            # Identificar linhas de título e subtítulos
            for row_idx, row in enumerate(df_resumo.values, start=1):
                cell_value_a = str(row[0]) if row[0] is not None and str(row[0]) != 'nan' else ''
                cell_value_b = str(row[1]) if len(row) > 1 and row[1] is not None and str(row[1]) != 'nan' else ''

                # Verificar se é título (está na lista de títulos E coluna B está vazia)
                if cell_value_a in titulos and not cell_value_b:
                    # Mesclar células A e B
                    worksheet.merge_cells(f'A{row_idx}:B{row_idx}')
                    # Aplicar negrito
                    worksheet[f'A{row_idx}'].font = Font(bold=True)

                # Verificar se é subtítulo (linha logo após título, com conteúdo em ambas as colunas)
                elif cell_value_a and cell_value_b and 'Total' not in cell_value_a:
                    # Verificar se a linha anterior era um título
                    if row_idx > 1:
                        prev_cell_value = str(df_resumo.values[row_idx - 2][0]) if df_resumo.values[row_idx - 2][0] is not None and str(df_resumo.values[row_idx - 2][0]) != 'nan' else ''
                        # Se a linha anterior é um título, esta é uma linha de cabeçalho de tabela
                        if prev_cell_value in titulos:
                            # Aplicar negrito nas células A e B da linha
                            worksheet[f'A{row_idx}'].font = Font(bold=True)
                            worksheet[f'B{row_idx}'].font = Font(bold=True)

        # Aba de votações
        if not df_votacoes.empty:
            df_votacoes.to_excel(writer, index=False, sheet_name='Votações CI')
            worksheet = writer.sheets['Votações CI']

            # Aplicar formato de data na coluna data_inicio
            if 'data_inicio' in df_votacoes.columns:
                col_idx = df_votacoes.columns.get_loc('data_inicio') + 1
                for row in range(2, len(df_votacoes) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = 'DD/MM/YYYY'

        # Aba de audiências públicas
        if not df_audiencias.empty:
            df_audiencias.to_excel(writer, index=False, sheet_name='Audiências Públicas CI')
            worksheet = writer.sheets['Audiências Públicas CI']

            # Aplicar formato de data na coluna data_inicio
            if 'data_inicio' in df_audiencias.columns:
                col_idx = df_audiencias.columns.get_loc('data_inicio') + 1
                for row in range(2, len(df_audiencias) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = 'DD/MM/YYYY'

        # Aba de partes das reuniões
        if not df_partes.empty:
            df_partes.to_excel(writer, index=False, sheet_name='Partes das Reuniões CI')
            worksheet = writer.sheets['Partes das Reuniões CI']

            # Aplicar formato de data na coluna data_inicio
            if 'data_inicio' in df_partes.columns:
                col_idx = df_partes.columns.get_loc('data_inicio') + 1
                for row in range(2, len(df_partes) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = 'DD/MM/YYYY'

    # Salvar também CSVs separados
    if not df_votacoes.empty:
        caminho_csv_votacoes = caminho_saida.replace('.xlsx', '_votacoes.csv')
        df_votacoes.to_csv(caminho_csv_votacoes, index=False, encoding='utf-8-sig')

    if not df_audiencias.empty:
        caminho_csv_audiencias = caminho_saida.replace('.xlsx', '_audiencias.csv')
        df_audiencias.to_csv(caminho_csv_audiencias, index=False, encoding='utf-8-sig')

    if not df_partes.empty:
        caminho_csv_partes = caminho_saida.replace('.xlsx', '_partes.csv')
        df_partes.to_csv(caminho_csv_partes, index=False, encoding='utf-8-sig')

    print(f"\n{'='*70}")
    print(f"Planilha consolidada gerada com sucesso!")
    print(f"Arquivo Excel: {caminho_saida}")
    print(f"{'='*70}")

    # Resumo de votações
    if not df_votacoes.empty:
        print(f"\nTotal de votações encontradas: {len(todas_votacoes)}")

        print("\nResumo por resultado:")
        if 'resultado' in df_votacoes.columns:
            resumo = df_votacoes['resultado'].value_counts()
            for resultado, qtd in resumo.items():
                print(f"  {resultado}: {qtd}")

        print("\nResumo por tipo de requerimento:")
        if 'tipo_requerimento' in df_votacoes.columns:
            df_req = df_votacoes[df_votacoes['tipo_requerimento'] != '']
            if not df_req.empty:
                resumo_req = df_req['tipo_requerimento'].value_counts()
                for tipo, qtd in resumo_req.items():
                    print(f"  {tipo}: {qtd}")
            else:
                print("  Nenhum requerimento encontrado")
    else:
        print("\nNenhuma votação encontrada na Comissão de Infraestrutura (CI)")

    # Resumo de audiências públicas
    print(f"\n{'='*70}")
    if not df_audiencias.empty:
        print(f"Total de audiências públicas encontradas: {len(todas_audiencias)}")
        print(f"\nResumo de audiências públicas:")
        print(f"  Realizadas: {df_audiencias[df_audiencias['realizada'] == 'true'].shape[0]}")
        print(f"  Não realizadas: {df_audiencias[df_audiencias['realizada'] == 'false'].shape[0]}")
        print(f"  Total de convidados: {df_audiencias['qtd_convidados'].sum()}")
        print(f"  Total de participantes: {df_audiencias['qtd_participantes'].sum()}")
    else:
        print("Nenhuma audiência pública encontrada na Comissão de Infraestrutura (CI)")
    print(f"{'='*70}")

    # Resumo de partes das reuniões
    print(f"\n{'='*70}")
    if not df_partes.empty:
        print(f"Total de partes de reuniões encontradas: {len(todas_partes)}")
        print(f"\nResumo por tipo de parte:")
        if 'tipo_parte' in df_partes.columns:
            resumo_tipo = df_partes['tipo_parte'].value_counts()
            for tipo, qtd in resumo_tipo.items():
                print(f"  {tipo}: {qtd}")

        print(f"\nResumo de reuniões conjuntas:")
        if 'reuniao_conjunta' in df_partes.columns:
            # Contar reuniões únicas (não partes)
            reunioes_unicas = df_partes[['titulo_reuniao', 'reuniao_conjunta']].drop_duplicates()
            qtd_conjuntas = reunioes_unicas[reunioes_unicas['reuniao_conjunta'] == 'Sim'].shape[0]
            qtd_nao_conjuntas = reunioes_unicas[reunioes_unicas['reuniao_conjunta'] == 'Não'].shape[0]
            print(f"  Reuniões conjuntas: {qtd_conjuntas}")
            print(f"  Reuniões não conjuntas: {qtd_nao_conjuntas}")

            # Mostrar comissões participantes se houver reuniões conjuntas
            if qtd_conjuntas > 0:
                comissoes_conjuntas = df_partes[df_partes['reuniao_conjunta'] == 'Sim']['comissoes_participantes'].unique()
                print(f"\nComissões que participaram de reuniões conjuntas com a CI:")
                for comissao in comissoes_conjuntas:
                    if comissao != '-':
                        print(f"  {comissao}")
    else:
        print("Nenhuma parte de reunião encontrada na Comissão de Infraestrutura (CI)")
    print(f"{'='*70}")


if __name__ == '__main__':
    # Obter o diretório do script
    pasta_atual = os.path.dirname(os.path.abspath(__file__))

    print("="*70)
    print("CONSOLIDADOR DE VOTAÇÕES - COMISSÃO DE INFRAESTRUTURA (CI)")
    print("="*70)

    # Solicitar ano ao usuário
    ano = solicitar_ano()

    print(f"\nAno selecionado: {ano}")

    # Verificar se os arquivos necessários existem
    print(f"\nVerificando arquivos na pasta {pasta_atual}...")
    tem_reunioes, tem_colegiados = verificar_arquivos_existentes(pasta_atual, ano)

    arquivos_faltando = []
    if not tem_reunioes:
        arquivos_faltando.append("arquivos de reuniões mensais")
    if not tem_colegiados:
        arquivos_faltando.append("lista de colegiados")

    if arquivos_faltando:
        print(f"\nArquivos faltando: {', '.join(arquivos_faltando)}")
        resposta = input("\nDeseja fazer o download dos arquivos faltantes? (s/n): ").strip().lower()

        if resposta in ['s', 'sim', 'y', 'yes']:
            sucesso = baixar_arquivos_ano(ano, pasta_atual)
            if not sucesso:
                print("\nAVISO: Alguns downloads falharam. Tentando processar com os arquivos disponíveis...")
        else:
            print("\nDownload cancelado pelo usuário.")
            print("O script tentará processar com os arquivos disponíveis na pasta.")
    else:
        print("Todos os arquivos necessários já existem na pasta.")

    # Processar arquivos e gerar planilha
    print(f"\n{'='*70}")
    print("INICIANDO PROCESSAMENTO")
    print(f"{'='*70}\n")

    gerar_planilha_consolidada(pasta_atual, ano)
